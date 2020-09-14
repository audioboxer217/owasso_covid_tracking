
import csv
from urllib.request import urlopen
import codecs
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta

wb_filename = '/Users/scott/Google Drive/Personal/Owasso_Covid.xlsx'
csv_url = 'https://storage.googleapis.com/ok-covid-gcs-public-download/oklahoma_cases_city.csv'
today = datetime.utcnow()

response = urlopen(csv_url)
cr = csv.reader(codecs.iterdecode(response, 'utf-8'))
wb = load_workbook(wb_filename)

ws = wb['imports']

prev_day = {
  ws['A2'].value: ws['B2'].value,
  ws['A3'].value: ws['B3'].value,
}

weekly_avg = {
  "OWASSO": [] ,
  "COLLINSVILLE": []  
}
dateWindow = today - timedelta(days=7)
for entry in ws.iter_rows(min_row=1):
  date = entry[7].value
  if date and date != 'Date' and date >= dateWindow:
    city = entry[0].value
    count = entry[2].value
    weekly_avg[city].append(count)

data = {}
for row in cr:
  if 'OWASSO' in row or 'COLLINSVILLE' in row:
    weekly_avg[row[0]].append(int(row[1]) - prev_day[row[0]])
    data[row[0]] = {
      "total": int(row[1]),
      "new": int(row[1]) - prev_day[row[0]],
      "average": sum(weekly_avg[row[0]]) / len(weekly_avg[row[0]]),
      "deaths": int(row[2]),
      "recovered": int(row[3]),
      "active": int(row[1]) - int(row[2]) - int(row[3]),
      "date": datetime.strptime(row[4], '%Y-%m-%d')
    }

ws.insert_rows(2,2)
ws['A2'] = "OWASSO"
ws['B2'] = data['OWASSO']['total']
ws['C2'] = data['OWASSO']['new']
ws['D2'] = data['OWASSO']['average']
ws['E2'] = data['OWASSO']['deaths']
ws['F2'] = data['OWASSO']['recovered']
ws['G2'] = data['OWASSO']['active']
ws['H2'] = data['OWASSO']['date']
ws['A3'] = "COLLINSVILLE"
ws['B3'] = data['COLLINSVILLE']['total']
ws['C3'] = data['COLLINSVILLE']['new']
ws['D3'] = data['COLLINSVILLE']['average']
ws['E3'] = data['COLLINSVILLE']['deaths']
ws['F3'] = data['COLLINSVILLE']['recovered']
ws['G3'] = data['COLLINSVILLE']['active']
ws['H3'] = data['COLLINSVILLE']['date']

wb.save(wb_filename)